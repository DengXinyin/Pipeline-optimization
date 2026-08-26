#!/usr/bin/env python3
"""Generate the annotated v2.88.2 WDL parameter workbook."""

from pathlib import Path
from datetime import datetime
import re

import WDL
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
WDL_PATH = ROOT / "metage_v2.88.2.wdl"
OUT_PATH = ROOT / "metage_v2.88.2_WDL参数说明.xlsx"
WORKFLOW = "metage_v2_88_2"


WORKFLOW_INPUTS = {
    "datapath": (
        "项目输入目录；必须包含 data.xlsx、project_info.json、report_no.txt。",
        "每次运行必须填写实际项目数据目录。data.xlsx 至少有 sample、comparison 两个工作表。",
        "必填（每次运行）", "所有模式",
    ),
    "rawdatapath": (
        "双端 FASTQ 原始数据目录。",
        "full/incremental 填可访问的 FASTQ 目录；reuse 建议保留原项目路径，便于本地规划器核对文件变化。",
        "full/incremental 必填", "full、incremental；reuse 仅规划器核对",
    ),
    "host": (
        "宿主序列去除类型。",
        "不去宿主填 none；按数据库支持情况可填 human 或 mouse。不要填写数据库未配置的名称。",
        "选填", "full、incremental",
    ),
    "mapdir": (
        "宏基因组公共数据库根目录。",
        "填写容器内可访问的完整 metagenome-DB 目录；通常使用 /public/nfs_data/public_file_data/metagenome-DB。",
        "必填（通常保留默认）", "所有产生注释/统计的模式",
    ),
    "binning": (
        "是否运行宏基因组分箱及 bins 统计。",
        "一般项目填 no；明确需要 MAG/binning 结果时填 yes。合法值仅 yes/no。",
        "选填", "主要用于 full；yes 时启用 bins 分支",
    ),
    "isbwa": (
        "历史兼容参数。当前 WDL 未用它控制 bwa_no。",
        "保留 yes 即可；不要依赖该参数跳过或启用比对。",
        "兼容参数", "当前版本不生效",
    ),
    "project_root": (
        "项目根目录，用于项目级 registry 和 incremental_data。",
        "填写稳定的项目根目录；流程约定读取 project_root/data/sample_registry.tsv 和 project_root/incremental_data。",
        "必填（尤其 reuse/incremental）", "所有模式；reuse/incremental 强依赖",
    ),
    "run_mode": (
        "运行方式：完整运行、增量运行或复用已有上游结果。",
        "首次/完整重跑填 full；新增样本填 incremental；仅删样本、改名、改分组填 reuse。",
        "必填（通常保留默认）", "合法值 full、incremental、reuse",
    ),
    "parent_workflow_dir": (
        "父流程 Cromwell Workflow UUID。WDL 会自动拼接固定绝对根目录。",
        "full 留空；首次增量通常填 full UUID，连续增量填最近一次成功且累计样本最完整的 incremental UUID；不填完整路径。",
        "incremental/reuse 必填", "run_mode != full",
    ),
    "use_kraken2": (
        "是否启用 Kraken2 物种注释。String 开关。",
        "一般填 no；启用时填 yes，并同时提供 kraken2_db。不要使用 JSON true/false。",
        "选填", "full/incremental 且值为 yes",
    ),
    "kraken2_db": (
        "Kraken2 数据库目录。",
        "use_kraken2=yes 时填写节点和容器均可访问的数据库目录；否则留空/null。",
        "条件必填", "use_kraken2=yes",
    ),
    "ref_sample": (
        "触发参考样本组装、参考比对和 SNP calling 的样本 ID。",
        "不做参考/SNP 分析留空；需要时填写 data.xlsx 中可识别的样本 ID，不是 fasta 路径。",
        "选填", "full/incremental 且非空",
    ),
}


WORKFLOW_OUTPUTS = {
    "respath": ("最终交付结果目录。", "始终输出"),
    "pdfFile": ("最终 PDF 报告。", "始终输出"),
    "docxpath": ("最终 Word/docx 报告。", "始终输出"),
    "jsonpath": ("前端结果索引 JSON 目录或文件。", "始终输出"),
    "reportNo": ("本次使用的报告编号文件。", "始终输出"),
    "infoFile": ("本次使用的 project_info.json。", "始终输出"),
    "kraken2_out": ("Kraken2 原始物种注释结果。", "use_kraken2=yes 且提供 kraken2_db"),
    "kraken2_tax_base_result": ("Kraken2 基础物种统计结果。", "Kraken2 注释分支运行时"),
    "kraken2_tax_diff_result": ("Kraken2 差异分析结果。", "use_kraken2=yes、full"),
    "ref_assembly_dir": ("参考样本组装结果目录。", "ref_sample 非空"),
    "ref_mapping_dir": ("参考序列比对结果目录。", "ref_sample 非空"),
    "snp_dir": ("SNP calling 结果目录。", "ref_sample 非空"),
}


TASK_DESCRIPTIONS = {
    "choose_plot_style": "生成全流程统一绘图样式配置",
    "prepare_registry_context": "检查并读取 reuse/incremental 的项目 registry 与增量目录",
    "check_input_with_raw": "校验三个输入文件、样本表、分组和 FASTQ 对应关系",
    "check_input_no_raw": "reuse 时校验三个输入文件和样本/分组信息",
    "kneaddata_no": "原始数据质控和可选宿主去除",
    "kraken2_anno": "Kraken2 物种注释",
    "kraken2_tax_base": "Kraken2 基础物种统计",
    "kraken2_tax_diff": "Kraken2 物种差异分析",
    "megahit_no": "MEGAHIT 宏基因组组装",
    "bins": "宏基因组分箱",
    "bins_drep": "bin 去冗余",
    "quant_classify": "bin 定量、分类与 blobology",
    "bins_stats": "bin 结果统计和可视化",
    "prodig_no": "Prodigal 基因预测",
    "bwa_no": "清洁 reads 回帖到基因集并定量",
    "tax_anno": "NR 等数据库物种注释",
    "func_anno": "eggNOG-mapper 功能注释",
    "deal_parameter": "优先定位父 Workflow 的累计 merged 状态，并兼容旧 full/cacheCopy",
    "merge_upstream_results": "合并上游、QC、基因集和 8 个数据库的历史累计/新增结果",
    "anno": "合并物种、功能与丰度，生成 GeneID 注释主目录",
    "VCA_anno": "VFDB、CycDB、ARG 注释",
    "MBQ_anno": "mobileOG、BacMet2、QS 注释",
    "COG_anno": "COG 注释",
    "MetaCyc_anno": "MetaCyc 注释",
    "apply_registry": "应用当前样本状态和显示名，生成最终分析 metadata",
    "tax_base": "物种基础统计与可视化",
    "func_base": "功能基础统计、KEGG 增强表及 GeneID 综合注释汇总",
    "tax_diff": "物种差异分析",
    "func_diff": "功能差异分析",
    "tax_unifrac": "物种四距离 Beta 分析",
    "func_unifrac": "功能四距离 Beta 分析",
    "coll_res_ana": "合并有差异分析、无 binning 的结果",
    "coll_res_ana_bins": "合并有差异分析且含 binning 的结果",
    "coll_res_NOana": "合并不做差异分析的结果",
    "res2json": "生成前端结果索引 JSON",
    "resFile": "整理最终结果并输出报告",
    "ref_assembly": "指定样本参考序列组装",
    "ref_mapping": "样本 reads 到参考序列的比对",
    "snp_calling": "参考比对后的 SNP 检测和统计",
    "update_registry": "流程成功后更新项目级样本 registry",
}


TASK_CONDITIONS = {
    "choose_plot_style": "始终运行",
    "prepare_registry_context": "run_mode != full",
    "check_input_with_raw": "run_mode=full 或 incremental",
    "check_input_no_raw": "run_mode=incremental 或 reuse",
    "kneaddata_no": "run_mode=full 或 incremental",
    "megahit_no": "run_mode=full 或 incremental",
    "prodig_no": "run_mode=full 或 incremental",
    "bwa_no": "run_mode=full 或 incremental",
    "tax_anno": "run_mode=full 或 incremental",
    "func_anno": "run_mode=full 或 incremental",
    "kraken2_anno": "full/incremental，use_kraken2=yes 且 kraken2_db 非空",
    "kraken2_tax_base": "Kraken2 注释分支运行时",
    "kraken2_tax_diff": "use_kraken2=yes、kraken2_db 非空、run_mode=full",
    "bins": "full/incremental 且 binning=yes",
    "bins_drep": "full/incremental 且 binning=yes",
    "quant_classify": "full/incremental 且 binning=yes",
    "bins_stats": "full/incremental 且 binning=yes",
    "ref_assembly": "full/incremental 且 ref_sample 非空",
    "ref_mapping": "ref_assembly 成功后",
    "snp_calling": "ref_mapping 成功后",
    "deal_parameter": "run_mode=incremental 或 reuse",
    "merge_upstream_results": "run_mode=incremental",
    "anno": "full 跑完整集；incremental 先跑新增集，合并后再重建累计 Annotation",
    "VCA_anno": "full 跑完整集；incremental 仅跑新增基因",
    "MBQ_anno": "full 跑完整集；incremental 仅跑新增基因",
    "COG_anno": "full 跑完整集；incremental 仅跑新增基因",
    "MetaCyc_anno": "full 跑完整集；incremental 仅跑新增基因",
    "apply_registry": "始终运行",
    "tax_base": "始终运行",
    "func_base": "始终运行",
    "tax_diff": "始终运行；物种四距离由上游注释和 taxonomy 自动生成",
    "func_diff": "始终运行",
    "tax_unifrac": "兼容旧版的未调用 task",
    "func_unifrac": "兼容旧版的未调用 task",
    "coll_res_ana": "binning=no",
    "coll_res_ana_bins": "binning=yes",
    "coll_res_NOana": "兼容旧版的未调用 task",
    "res2json": "始终运行",
    "resFile": "始终运行",
    "update_registry": "resFile 成功后",
}


PARAM_MEANINGS = {
    "workflow_dir": "父 workflow 的完整执行目录（由固定根路径加 UUID 自动生成）",
    "dataDir": "待校验的输入数据目录",
    "fastq_dir": "FASTQ 原始数据目录",
    "bust_cache": "registry 内容摘要，用于使 metadata 检查缓存失效",
    "allow_extra_fastq": "是否允许 FASTQ 目录包含本次增量表之外的样本",
    "allow_empty_comparison": "是否允许增量 data.xlsx 的 comparison 为空",
    "project_root": "项目根目录",
    "require_incremental_data": "是否强制检查 incremental_data 目录",
    "datapath": "上游生成的标准化 metadata 目录",
    "rawdatapath": "FASTQ 原始数据目录",
    "host": "宿主去除类型",
    "mapdir": "宏基因组公共数据库根目录",
    "checkDir": "输入检查结果目录",
    "keep_clean_reads": "是否保留清洁 reads 供 Kraken2 使用",
    "clean_dir": "质控后的 clean reads 目录",
    "cleandir": "质控后的 clean reads 目录",
    "dehost_dir": "去宿主后的 reads 目录",
    "dohost_dir": "去宿主结果目录",
    "megahit": "MEGAHIT 组装结果目录",
    "binsDir": "binning 输出目录",
    "drepDir": "dRep 去冗余结果目录",
    "classfiDir": "bin 分类结果目录",
    "quantDir": "bin 定量结果目录",
    "blobologyDir": "blobology 结果目录",
    "prodigal": "Prodigal 基因预测结果目录",
    "bowtie": "reads 回帖/基因丰度结果目录",
    "tax_Annotation": "物种注释结果目录",
    "func_Annotation": "功能注释结果目录",
    "Annotation": "综合 GeneID 注释目录",
    "ARGdir": "抗性基因注释目录",
    "CycDB": "元素循环注释目录",
    "VFDB": "毒力因子注释目录",
    "mobileOGs": "可移动遗传元件注释目录",
    "BacMet2": "金属/生物杀灭剂抗性注释目录",
    "QS": "群体感应注释目录",
    "COG": "COG 注释目录",
    "MetaCyc": "MetaCyc 注释目录",
    "preResdir": "前一基础分析结果目录",
    "funcBase": "功能差异/UniFrac 使用的临时丰度表目录",
    "plot_style": "流程统一绘图样式 JSON",
    "task_overrides_json": "单个绘图 task 的样式覆盖 JSON",
    "registry_tsv": "父项目样本 registry 路径",
    "registry_md5": "registry 文件 MD5，用于缓存控制",
    "registry_tsv_path": "要更新的项目级 registry 路径",
    "project_info": "标准化 project_info.json",
    "workflow_success_marker": "证明最终结果成功生成的文件",
    "display_name_map": "内部样本 ID 到报告显示名的映射表",
    "qc_cleandir": "供结果整理使用的清洁 reads 目录",
    "analyse": "是否进行差异分析",
    "binning": "是否包含 binning 分支",
    "res_dir": "待整理或索引的结果目录",
    "report_no": "报告编号文件",
    "projectinfo": "项目信息 JSON",
    "kraken2_db": "Kraken2 数据库目录",
    "kraken2_out": "Kraken2 注释结果目录",
    "threads": "任务线程数",
    "ref_sample": "参考组装样本 ID",
    "ref_fasta": "参考序列 FASTA",
    "bamdir": "参考比对 BAM 目录",
    "tax_tree": "物种 Newick 系统发育树",
    "func_tree": "功能 Newick 特征树",
    "resdir": "物种基础结果目录",
}


OUTPUT_MEANINGS = {
    "clean_dir": "合并后的 clean reads 目录，或父流程 clean reads 目录",
    "qc_result": "合并后的质控统计结果目录",
    "tax_annotation": "合并后的物种注释目录",
    "func_annotation": "合并后的功能注释目录",
    "megahitdir": "父流程 MEGAHIT 组装结果目录",
    "prodigdir": "父流程 Prodigal 基因预测结果目录",
    "bwadir": "父流程 reads 回帖/基因定量结果目录",
    "func_annodir": "父流程功能注释原始目录",
    "anno_dir": "父流程综合 GeneID 注释目录",
    "tax_annodir": "父流程物种注释原始目录",
    "ARGsdir": "父流程 ARG 注释目录",
    "CycDBdir": "父流程 CycDB 注释目录",
    "VFDBdir": "父流程 VFDB 注释目录",
    "BacMet2_annodir": "父流程 BacMet2 注释目录",
    "QS_annodir": "父流程 QS 注释目录",
    "mobileOG_annodir": "父流程 mobileOG 注释目录",
    "COGdir": "父流程 COG 注释目录",
    "MetaCycdir": "父流程 MetaCyc 注释目录",
    "kneaddatadir": "父流程质控统计 Result 目录",
    "bins_stats_dir": "父流程 bins 统计结果目录（当前 WDL 表达式沿用 kneaddata Result）",
    "result": "标准化输入检查结果目录",
    "metadata_files": "标准化 metadata 文件列表",
    "project_info": "标准化 project_info.json",
    "report_no": "标准化报告编号文件",
    "plot_style": "统一绘图样式 JSON",
    "registry_md5": "registry 内容 MD5",
    "registry_tsv": "registry 路径字符串",
    "incremental_datapath": "增量数据目录路径字符串",
    "new_datapath": "应用当前样本状态后的 metadata 目录",
    "display_name_map": "内部 ID 与显示名映射表",
    "cleandir": "clean reads 目录",
    "Result": "当前 task 的结果目录",
    "dohost_dir": "去宿主结果目录",
    "megahit": "MEGAHIT 组装目录",
    "binsDir": "binning 结果目录",
    "drepDir": "dRep 结果目录",
    "classfiDir": "bin 分类结果目录",
    "quantDir": "bin 定量结果目录",
    "blobologyDir": "blobology 结果目录",
    "prodigal": "基因预测结果目录",
    "bowtie": "基因定量/回帖结果目录",
    "tax_Annotation": "物种注释目录",
    "func_Annotation": "功能注释目录",
    "Annotation": "综合注释目录",
    "VFDB": "VFDB 注释目录",
    "CycDB": "CycDB 注释目录",
    "ARGdir": "ARG 注释目录",
    "mobileOGs": "mobileOG 注释目录",
    "BacMet2": "BacMet2 注释目录",
    "QS": "QS 注释目录",
    "COG": "COG 注释目录",
    "MetaCyc": "MetaCyc 注释目录",
    "funcBase": "功能分析临时表目录",
    "jsonFile": "前端结果索引 JSON",
    "respath": "最终结果目录",
    "PDFpath": "最终 PDF 报告",
    "docxpath": "最终 Word 报告",
    "reportNOdir": "报告编号文件",
    "kraken2_out": "Kraken2 原始结果目录",
    "ref_assembly_dir": "参考组装结果目录",
    "ref_fasta": "参考 FASTA",
    "ref_mapping_dir": "参考比对结果目录",
    "snp_dir": "SNP 结果目录",
    "tax_unifrac_out": "物种四距离 Beta 结果目录",
    "func_unifrac_out": "功能四距离 Beta 结果目录",
    "updated_registry_tsv_path": "更新后的 registry 路径",
}


def expr_text(expr):
    if expr is None:
        return ""
    return str(expr)


def task_param_meaning(name):
    if name in PARAM_MEANINGS:
        return PARAM_MEANINGS[name]
    if re.fullmatch(r"old_.+", name):
        return "父流程的%s" % PARAM_MEANINGS.get(name[4:], name[4:])
    if re.fullmatch(r"new_.+", name):
        return "本次新增样本的%s" % PARAM_MEANINGS.get(name[4:], name[4:])
    if re.fullmatch(r"Res[1-6]", name):
        labels = {
            "Res1": "质控结果", "Res2": "物种基础结果", "Res3": "功能基础结果",
            "Res4": "物种差异结果", "Res5": "功能差异结果", "Res6": "binning 统计结果",
        }
        return labels[name]
    if name.startswith("global_"):
        return "全局绘图%s" % name.replace("global_", "")
    for prefix, label in (
        ("title_", "图标题"), ("axis_title_", "坐标轴标题"),
        ("axis_text_", "坐标轴文字"), ("legend_title_", "图例标题"),
        ("legend_text_", "图例文字"), ("label_", "数据标签"),
        ("legend_", "图例"),
    ):
        if name.startswith(prefix):
            return "%s的%s设置" % (label, name[len(prefix):])
    if name == "group_palette":
        return "分组颜色列表"
    return "WDL 内部任务参数：%s" % name


def walk_calls(nodes, result):
    for node in nodes:
        if type(node).__name__ == "Call":
            result[node.name] = node
        body = getattr(node, "body", None)
        if isinstance(body, list):
            walk_calls(body, result)


def set_sheet_table(ws, headers, rows, widths, table_name):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="D9E2F3"))
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    if rows:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)


def runtime_value(task, name):
    return expr_text(task.runtime.get(name)) if name in task.runtime else ""


def main():
    doc = WDL.load(str(WDL_PATH))
    workflow = doc.workflow
    workflow_decls = [node for node in workflow.body if type(node).__name__ == "Decl"]
    calls = {}
    walk_calls(workflow.body, calls)

    wb = Workbook()
    intro = wb.active
    intro.title = "使用说明"
    intro.sheet_view.showGridLines = False
    intro["A1"] = "metage_v2.88.2 WDL 参数说明"
    intro["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    intro["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    intro.merge_cells("A1:F1")
    intro["A2"] = "生成时间"
    intro["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    intro["A3"] = "WDL 文件"
    intro["B3"] = WDL_PATH.name
    intro["A4"] = "workflow"
    intro["B4"] = workflow.name
    intro["A5"] = "固定父流程根目录"
    intro["B5"] = "/cephfs_data/genostack_v3/genostack_cromwell/cromwell-executions/metage_v2_88_2/<Workflow UUID>"
    intro["A7"] = "填写原则"
    intro["B7"] = "分析者主要填写“工作流输入”工作表中的参数；Task 输入绝大多数由 WDL 自动传递，不应手工填写。"
    intro["A8"] = "三个必需文件"
    intro["B8"] = "datapath 目录必须同时包含 data.xlsx、project_info.json、report_no.txt。"
    intro["A9"] = "字符串开关"
    intro["B9"] = "binning、use_kraken2 使用 yes/no 字符串；物种 Beta 默认输出四种距离，无需提交分析开关或树路径。"
    intro["A10"] = "父流程参数"
    intro["B10"] = "parent_workflow_dir 只填成功父流程 Workflow UUID；WDL 自动拼接固定绝对路径。"
    intro["A11"] = "Task ID 说明"
    intro["B11"] = "平台若把 Workflow UUID 显示为 Task ID，应确认该值就是 Cromwell execution 根目录下的 UUID 文件夹名。"
    intro["A13"] = "工作表"
    intro["B13"] = "工作流输入：用户配置；Task输入参数：内部传递与高级线程项；工作流输出/Task输出参数：结果；Task运行资源：CPU、内存和镜像。"
    intro.column_dimensions["A"].width = 22
    intro.column_dimensions["B"].width = 115
    for row in intro.iter_rows(min_row=2, max_row=13, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in (7, 8, 9, 10, 11, 13):
        intro.cell(row, 1).font = Font(bold=True, color="1F4E78")

    input_rows = []
    for decl in workflow_decls:
        meaning, advice, required, condition = WORKFLOW_INPUTS[decl.name]
        input_rows.append([
            "%s.%s" % (WORKFLOW, decl.name), decl.name, str(decl.type),
            expr_text(decl.expr) or "null/未设置", required, meaning, condition, advice,
        ])
    for task_name, parameter, default, meaning, condition, advice in (
        ("kraken2_anno", "threads", "16", "Kraken2 注释线程数", "Kraken2 注释分支运行时", "一般保留 16；按节点 CPU 配额调整。"),
        ("kraken2_tax_diff", "threads", "8", "Kraken2 差异分析线程数", "Kraken2 差异分支运行时", "一般保留 8；按节点 CPU 配额调整。"),
    ):
        input_rows.append([
            "%s.%s.%s" % (WORKFLOW, task_name, parameter),
            "%s.%s" % (task_name, parameter), "Int", default,
            "高级选填", meaning, condition, advice,
        ])
    ws = wb.create_sheet("工作流输入")
    set_sheet_table(
        ws,
        ["平台完整参数名", "短参数名", "类型", "WDL默认值", "填写要求", "含义", "生效条件", "建议选择/填写"],
        input_rows, [48, 28, 14, 50, 22, 45, 38, 70], "WorkflowInputs",
    )
    for row in range(2, ws.max_row + 1):
        if "必填" in str(ws.cell(row, 5).value):
            ws.cell(row, 5).fill = PatternFill("solid", fgColor="FFF2CC")

    task_input_rows = []
    for task in doc.tasks:
        call = calls.get(task.name)
        declarations = task.postinputs or task.inputs or []
        for decl in declarations:
            default = expr_text(decl.expr) or "无"
            if call and decl.name in call.inputs:
                source = expr_text(call.inputs[decl.name])
                user = "否（WDL 自动传递）"
                advice = "无需填写；由调用表达式自动传入。"
            elif decl.expr is not None:
                source = "task 默认值"
                if task.name in {"kraken2_anno", "kraken2_tax_diff"} and decl.name == "threads":
                    user = "可选（高级参数）"
                    advice = "通常保留默认值；仅按节点资源调整。"
                else:
                    user = "否（使用 task 默认值）"
                    advice = "无需填写；保留 task 默认值。"
            else:
                source = "由条件调用自动绑定"
                user = "否（流程内部参数）"
                advice = "不要在平台手工填写；检查对应上游 task 是否运行。"
            task_input_rows.append([
                task.name, TASK_DESCRIPTIONS.get(task.name, task.name), decl.name,
                str(decl.type), default, user, task_param_meaning(decl.name),
                TASK_CONDITIONS.get(task.name, "按 WDL 调用条件"), source, advice,
            ])
    ws = wb.create_sheet("Task输入参数")
    set_sheet_table(
        ws,
        ["Task", "Task作用", "参数名", "类型", "默认值", "用户是否填写", "含义", "Task运行条件", "WDL传入来源/表达式", "建议"],
        task_input_rows, [28, 42, 28, 16, 30, 24, 42, 42, 75, 48], "TaskInputs",
    )
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, 6).value).startswith("可选"):
            ws.cell(row, 6).fill = PatternFill("solid", fgColor="E2F0D9")

    workflow_output_rows = []
    for decl in workflow.outputs:
        meaning, condition = WORKFLOW_OUTPUTS.get(decl.name, (decl.name, "按 WDL 条件"))
        workflow_output_rows.append([
            "%s.%s" % (WORKFLOW, decl.name), decl.name, str(decl.type),
            expr_text(decl.expr), meaning, condition,
        ])
    ws = wb.create_sheet("工作流输出")
    set_sheet_table(
        ws, ["平台完整输出名", "输出名", "类型", "来源表达式", "含义", "存在条件"],
        workflow_output_rows, [48, 30, 14, 48, 55, 48], "WorkflowOutputs",
    )

    task_output_rows = []
    for task in doc.tasks:
        for decl in task.outputs:
            task_output_rows.append([
                task.name, TASK_DESCRIPTIONS.get(task.name, task.name), decl.name,
                str(decl.type), expr_text(decl.expr),
                OUTPUT_MEANINGS.get(decl.name, "当前 task 输出：%s" % decl.name),
                TASK_CONDITIONS.get(task.name, "按 WDL 调用条件"),
            ])
    ws = wb.create_sheet("Task输出参数")
    set_sheet_table(
        ws, ["Task", "Task作用", "输出名", "类型", "WDL路径/表达式", "含义", "存在条件"],
        task_output_rows, [28, 44, 32, 16, 72, 48, 48], "TaskOutputs",
    )

    resource_rows = []
    for task in doc.tasks:
        resource_rows.append([
            task.name, TASK_DESCRIPTIONS.get(task.name, task.name),
            runtime_value(task, "cpu"), runtime_value(task, "memory"),
            runtime_value(task, "docker"), TASK_CONDITIONS.get(task.name, "按 WDL 调用条件"),
        ])
    ws = wb.create_sheet("Task运行资源")
    set_sheet_table(
        ws, ["Task", "Task作用", "CPU", "内存", "Docker镜像", "运行条件"],
        resource_rows, [28, 48, 12, 18, 60, 52], "TaskRuntime",
    )

    wb.properties.title = "metagenome v2.88.2 WDL 参数说明"
    wb.properties.subject = "WDL workflow/task 输入输出、运行条件及填写建议"
    wb.properties.creator = "Codex"
    wb.save(OUT_PATH)

    # Re-open to ensure the file is structurally readable.
    checked = load_workbook(OUT_PATH, read_only=True, data_only=False)
    assert checked.sheetnames == [
        "使用说明", "工作流输入", "Task输入参数", "工作流输出", "Task输出参数", "Task运行资源"
    ]
    checked.close()
    print(OUT_PATH)
    print("workflow_inputs=%d task_inputs=%d workflow_outputs=%d task_outputs=%d tasks=%d" % (
        len(input_rows), len(task_input_rows), len(workflow_output_rows),
        len(task_output_rows), len(resource_rows),
    ))


if __name__ == "__main__":
    main()
