#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据 display_name_map.tsv，把结果目录中的 internal_id 替换为 display_name。

替换范围：
- 文本文件（tsv/csv/txt/md/json）内容中的 whole-word 匹配
- Excel 文件（xlsx）所有单元格中的 whole-word 匹配
- 文件/目录名中的 whole-word 匹配（谨慎执行）

注意：图片中的文字无法替换；报告 docx/pdf 建议在 get_report_update.py 中单独处理。
"""

import argparse
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path

import pandas as pd


# 结果目录中按样本组织的父目录（这些目录下会按样本名创建子目录）
SAMPLE_SUBDIR_PARENTS = {"1-data_quality", "2-Assembly", "3-GenePredict"}
# 这些目录不是样本目录，不应被删除
PROTECTED_DIRS = {"contig_length_files", "gene_length_files"}


def load_map(map_tsv):
    df = pd.read_csv(map_tsv, sep="\t", dtype=str)
    if "internal_id" not in df.columns or "display_name" not in df.columns:
        raise ValueError(f"{map_tsv} 必须包含 internal_id 和 display_name 列")
    mapping = {}
    kept_samples = set()
    for _, r in df.iterrows():
        iid = str(r["internal_id"]) if pd.notna(r["internal_id"]) else ""
        display = str(r["display_name"]) if pd.notna(r["display_name"]) else ""
        if iid:
            kept_samples.add(iid)
        if iid and display and iid != display:
            mapping[iid] = display
    return mapping, kept_samples


def _build_pattern(iid):
    """构建 whole-word 匹配正则，internal_id 中的特殊字符会被转义。"""
    return re.compile(r"(?<!\w)" + re.escape(iid) + r"(?!\w)")


def replace_text_in_file(file_path, mapping, encoding="utf-8"):
    """替换单个文本文件中的 internal_id。"""
    try:
        with open(file_path, "r", encoding=encoding, errors="ignore") as f:
            content = f.read()
    except Exception as e:
        print(f"  跳过读取失败: {file_path}: {e}")
        return False

    new_content = content
    for iid, display in mapping.items():
        new_content = _build_pattern(iid).sub(display, new_content)

    if new_content == content:
        return False

    try:
        with open(file_path, "w", encoding=encoding) as f:
            f.write(new_content)
        print(f"  已替换: {file_path}")
        return True
    except Exception as e:
        print(f"  写入失败: {file_path}: {e}")
        return False


def rewrite_excel(file_path, mapping):
    """替换 xlsx 文件中所有单元格的 internal_id。"""
    try:
        import openpyxl
    except ImportError:
        print(f"  未安装 openpyxl，跳过 Excel 处理: {file_path}")
        return False

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"  读取 Excel 失败: {file_path}: {e}")
        return False

    changed = False
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str):
                    new_val = cell.value
                    for iid, display in mapping.items():
                        new_val = _build_pattern(iid).sub(display, new_val)
                    if new_val != cell.value:
                        cell.value = new_val
                        changed = True

    if not changed:
        return False

    try:
        wb.save(file_path)
        print(f"  已替换 Excel: {file_path}")
        return True
    except Exception as e:
        print(f"  保存 Excel 失败: {file_path}: {e}")
        return False


def replot_qc(res_dir, table_dir, data_dir, host):
    """使用 display_name_map 重新绘制 kneaddata 阶段生成的 QC 图。

    kneaddata_no 在 apply_registry 之前运行，使用的原始 metadatadir 没有
    display_name_map.tsv，因此 error_rate/ATGC/reads_quality_summary 图标题仍
    为 internal_id。这里利用已生成的 QC table 和 new_metadatadir 中的映射表
    重新生成这三类图，覆盖原有文件。
    """
    rscript = "/root/anaconda3/envs/r/bin/Rscript"
    scripts_dir = "/root/microbiome/microbiome/metage_v2.88.2"

    if not os.path.isdir(table_dir):
        print(f"[replot_qc] 跳过：QC table 目录不存在 {table_dir}")
        return
    if not os.path.isdir(data_dir):
        print(f"[replot_qc] 跳过：data_dir 不存在 {data_dir}")
        return

    # Result_update 的层级结构为 Result_update/Result/group*/...，
    # R 脚本需要写入 Result 子目录，否则会在 Result_update 下新建多余的 group1。
    qc_res_dir = os.path.join(res_dir, "Result")
    if not os.path.isdir(qc_res_dir):
        print(f"[replot_qc] 跳过：Result 子目录不存在 {qc_res_dir}")
        return

    print(f"[replot_qc] 使用 QC table {table_dir} 和 data_dir {data_dir} 重新绘制 QC 图")

    # error_rate 与 ATGC_content 参数相同
    for rscript_name in ("error_rate_update.R", "atgc_content_update.R"):
        cmd = [
            rscript,
            os.path.join(scripts_dir, rscript_name),
            table_dir,
            data_dir,
            qc_res_dir,
        ]
        print(f"[replot_qc] {' '.join(cmd)}")
        subprocess.run(cmd, check=True)

    # data_composition_bar (reads_quality_summary) 需要 host 参数
    cmd = [
        rscript,
        os.path.join(scripts_dir, "data_composition_bar_update.R"),
        table_dir,
        data_dir,
        qc_res_dir,
        host,
    ]
    print(f"[replot_qc] {' '.join(cmd)}")
    subprocess.run(cmd, check=True)

    print("[replot_qc] QC 图重绘完成")


def cleanup_deleted_samples(res_dir, kept_samples, mapping=None):
    """当用户从 registry 中删除样本后，清理 Result_update 中残留的已删除样本目录和 Excel 行。

    清理范围：
    - Result_update/Result/group*/{1-data_quality,2-Assembly,3-GenePredict}/ 下
      不在保留样本集合中的子目录（保留 internal_id 或 display_name 均视为有效）。
    - Result_update/Result/group*/1-data_quality/data_quality.xlsx 中
      Sample_name 列不在保留样本集合中的行。

    注意：目录名可能仍是 internal_id（首次运行），也可能已被改为 display_name（前次改名后
    再次删除）。因此保留集合同时包含 kept_samples 和 mapping.values()。
    """
    res_dir = Path(res_dir)
    if not res_dir.exists() or not kept_samples:
        return

    kept_internal = set(kept_samples)
    kept_display = set()
    if mapping:
        kept_display = set(v for v in mapping.values() if v)
    kept_all = kept_internal | kept_display

    # 1. 清理按样本组织的子目录
    deleted_dirs = 0
    for group_dir in sorted(res_dir.glob("Result/group*")):
        if not group_dir.is_dir():
            continue
        for parent_name in SAMPLE_SUBDIR_PARENTS:
            parent_dir = group_dir / parent_name
            if not parent_dir.is_dir():
                continue
            for item in sorted(parent_dir.iterdir()):
                if not item.is_dir() or item.name in PROTECTED_DIRS:
                    continue
                if item.name not in kept_all:
                    try:
                        shutil.rmtree(item)
                        print(f"  删除残留样本目录: {item}")
                        deleted_dirs += 1
                    except Exception as e:
                        print(f"  删除目录失败: {item}: {e}")

    # 2. 清理 data_quality.xlsx 中已删除样本的行
    deleted_rows = 0
    for dq_xlsx in sorted(res_dir.glob("Result/group*/1-data_quality/data_quality.xlsx")):
        try:
            import openpyxl
        except ImportError:
            print(f"  未安装 openpyxl，跳过 Excel 清理: {dq_xlsx}")
            continue
        try:
            wb = openpyxl.load_workbook(dq_xlsx)
        except Exception as e:
            print(f"  读取 Excel 失败，跳过清理: {dq_xlsx}: {e}")
            continue

        ws = wb.active
        if ws is None or ws.max_row < 1:
            continue
        header = [cell.value for cell in ws[1]]
        sample_col = None
        for i, h in enumerate(header):
            if h and str(h).strip().lower() in ("sample_name", "sample-name", "sample", "sample_id", "sample id"):
                sample_col = i
                break
        if sample_col is None:
            print(f"  未找到 Sample_name 列，跳过 Excel 清理: {dq_xlsx}")
            continue

        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            val = row[sample_col].value
            if val is None:
                continue
            val = str(val).strip()
            if val not in kept_all:
                rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            try:
                ws.delete_rows(row_idx)
                deleted_rows += 1
            except Exception as e:
                print(f"  删除 Excel 行失败: {dq_xlsx} 行 {row_idx}: {e}")

        if rows_to_delete:
            try:
                wb.save(dq_xlsx)
                print(f"  已清理 Excel 中 {len(rows_to_delete)} 行已删除样本: {dq_xlsx}")
            except Exception as e:
                print(f"  保存 Excel 失败: {dq_xlsx}: {e}")

    if deleted_dirs or deleted_rows:
        print(f"[cleanup_deleted_samples] 共删除 {deleted_dirs} 个残留目录，{deleted_rows} 行 Excel 记录")


def rewrite_directory(res_dir, mapping):
    """遍历目录，替换文本/Excel 内容，并重命名文件/目录。"""
    res_dir = Path(res_dir)
    if not res_dir.exists():
        raise FileNotFoundError(f"结果目录不存在: {res_dir}")

    text_exts = {".tsv", ".csv", ".txt", ".md", ".json", ".html", ".xml"}
    excel_exts = {".xlsx"}

    # 1. 替换文件内容
    for file_path in sorted(res_dir.rglob("*")):
        if not file_path.is_file():
            continue
        ext = file_path.suffix.lower()
        if ext in text_exts:
            replace_text_in_file(file_path, mapping)
        elif ext in excel_exts:
            rewrite_excel(file_path, mapping)

    # 2. 重命名文件（先处理深层文件）
    for file_path in sorted(res_dir.rglob("*"), reverse=True):
        if not file_path.is_file():
            continue
        new_name = file_path.name
        for iid, display in mapping.items():
            new_name = _build_pattern(iid).sub(display, new_name)
        if new_name != file_path.name:
            new_path = file_path.with_name(new_name)
            try:
                file_path.rename(new_path)
                print(f"  重命名文件: {file_path} -> {new_path}")
            except Exception as e:
                print(f"  重命名失败: {file_path}: {e}")

    # 3. 重命名目录（先处理深层目录）
    for dir_path in sorted(res_dir.rglob("*"), reverse=True):
        if not dir_path.is_dir():
            continue
        new_name = dir_path.name
        for iid, display in mapping.items():
            new_name = _build_pattern(iid).sub(display, new_name)
        if new_name != dir_path.name:
            new_path = dir_path.with_name(new_name)
            try:
                dir_path.rename(new_path)
                print(f"  重命名目录: {dir_path} -> {new_path}")
            except Exception as e:
                print(f"  重命名失败: {dir_path}: {e}")


def main():
    parser = argparse.ArgumentParser(description="根据 display_name_map 替换结果目录中的样本展示名")
    parser.add_argument("--res_dir", required=True, help="结果目录（如 Result_update）")
    parser.add_argument("--map", required=True, help="display_name_map.tsv 路径")
    parser.add_argument("--qc-table-dir", dest="qc_table_dir", default=None,
                        help="kneaddata cleandata/table 路径，用于重绘 QC 图")
    parser.add_argument("--qc-data-dir", dest="qc_data_dir", default=None,
                        help="new_metadatadir 路径，用于读取 display_name_map.tsv")
    parser.add_argument("--host", default="none",
                        help="host 类型，传递给 data_composition_bar_update.R")
    args = parser.parse_args()

    mapping, kept_samples = load_map(args.map)
    if not kept_samples:
        print("display_name_map 为空或无需处理，退出")
        sys.exit(0)

    print(f"加载 display_name_map，保留样本 {len(kept_samples)} 个，需替换映射 {len(mapping)} 条:")
    for iid, display in sorted(mapping.items()):
        print(f"  {iid} -> {display}")

    # 在文本/目录替换前先重绘 QC 图，使新图使用 display_name，随后目录重命名生效
    if mapping and args.qc_table_dir and args.qc_data_dir:
        replot_qc(args.res_dir, args.qc_table_dir, args.qc_data_dir, args.host)

    # 清理用户从 registry 中删除的样本所残留的目录和 Excel 行
    cleanup_deleted_samples(args.res_dir, kept_samples, mapping)

    if mapping:
        rewrite_directory(args.res_dir, mapping)
        print("完成 display_name 替换")
    else:
        print("无需替换 display_name")


if __name__ == "__main__":
    main()
